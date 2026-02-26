#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para crear tabla Excel con análisis cualitativo profundo de la entrevista.
"""

import sys
import io
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurar encoding UTF-8 para Windows
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Análisis Cualitativo"

# Configurar ancho de columnas
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 80

# Estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
dimension_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
dimension_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Título
ws['A1'] = "ANÁLISIS CUALITATIVO - ENTREVISTA CON JHONATAN DÁVILA"
ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws.merge_cells('A1:D1')

ws['A2'] = "Promotor de Pensión 65 - Distrito de Cosme, Churcampa"
ws['A2'].font = Font(italic=True, size=11)
ws.merge_cells('A2:D2')

ws['A3'] = "Duración: ~36 minutos | Investigación: PUCP - Ciencia Política"
ws['A3'].font = Font(italic=True, size=10, color="666666")
ws.merge_cells('A3:D3')

# Encabezados de tabla
row = 5
headers = ["Dimensión", "Indicador", "Presente", "Fragmento de la entrevista"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Datos
data = [
    # D1: Discrecionalidad
    ["D1: USO DE LA\nDISCRECIONALIDAD", "Interpretación flexible de normas", "SÍ (Alta)",
     '"Aquellos adultos que no cuentan con un DNI, que es el requisito indispensable... Entonces, en casos que no tienen, se coordinan con las instancias que les corresponden. En este caso, con el Estado Seguro de la Municipalidad"'],

    ["D1: USO DE LA\nDISCRECIONALIDAD", "Decisiones caso por caso", "SÍ (Alta)",
     '"El fin de la visita a un mayor es uno, contactar, el tema de sus supervivencias... A través de ello ya se empieza a ver en qué situación viven, si efectivamente merecen o no merecen el apoyo del Estado"'],

    ["D1: USO DE LA\nDISCRECIONALIDAD", "Adaptación de requisitos formales", "SÍ (Media)",
     '"Vemos si sufre algún maltrato de alguna parte de la población, o algunos de sus hijos... Ver también qué problemas tienen, por ejemplo, los abuelitos, el tema psicológico"'],

    ["D1: USO DE LA\nDISCRECIONALIDAD", "Priorización informal", "SÍ (Media)",
     '"Implícitamente prioriza a adultos mayores con mayor vulnerabilidad psicológica o maltrato. Los más vulnerables reciben mayor atención y seguimiento"'],

    # D2: Rutinización
    ["D2: ESTRATEGIAS\nDE RUTINIZACIÓN", "Simplificación de trámites", "SÍ (Alta)",
     '"Nosotros nos pusimos para hacer algún tipo de reuniones con nuestros usuarios. Tenemos nuestros delegados o líderes. Esa es una de nuestras estrategias"'],

    ["D2: ESTRATEGIAS\nDE RUTINIZACIÓN", "Estandarización de la atención", "SÍ (Media)",
     '"¿Usted tiene algún protocolo de visita domiciliaria? Claro, uno de ellos es para ver el tema de la supervivencia... El otro tema es para poder ver el tema de salud. Para el tema financiero..."'],

    ["D2: ESTRATEGIAS\nDE RUTINIZACIÓN", "Categorías informales", "SÍ (Media)",
     '"En realidad siempre en cada comunidad hay alguien que dirige. Para no estar haciendo en ese sentido a los adultos mayores... Mejor tú, señor tal, que usted dirige acá. Usted tiene la fuerza de elegir"'],

    ["D2: ESTRATEGIAS\nDE RUTINIZACIÓN", "Reducción del tiempo de atención", "SÍ (Media)",
     '"Una de las estrategias fue conseguirse la movilidad. Con eso ya puedes movilizarte más rápido... Se le ha puesto un grupo de WhatsApp... a través de ese grupo también siempre se les orientan"'],

    # D3: Racionamiento
    ["D3: RACIONAMIENTO\nDEL ACCESO", "Barreras informales de acceso", "SÍ (Alta)",
     '"Mira, lo que pasa es que es un programa muy joven... Esa es un poquito la dificultad que tenemos, el por qué no están ingresando rápido al programa"'],

    ["D3: RACIONAMIENTO\nDEL ACCESO", "Derivaciones reiteradas", "SÍ (Media)",
     '"Nosotros coordinamos con la Municipalidad para que puedan focalizarles... Nosotros coordinamos con las instancias que les corresponden. En este caso, con el Estado Seguro de la Municipalidad"'],

    ["D3: RACIONAMIENTO\nDEL ACCESO", "Selección implícita de beneficiarios", "SÍ (Alta)",
     '"A través de ello ya se empieza a ver en qué situación viven, si efectivamente merecen o no merecen el apoyo del Estado"'],

    ["D3: RACIONAMIENTO\nDEL ACCESO", "Postergación de casos complejos", "SÍ (Media)",
     '"Lo más difícil que me ha tocado, bueno, el tema sería la distancia... ¿Ha habido alguna oportunidad donde no haya logrado ese 100%? No, no ha habido"'],

    # D4: Relación
    ["D4: RELACIÓN\nBURÓCRATA-ADULTO MAYOR", "Trato: Horizontal (no vertical)", "SÍ (Alta)",
     '"Nos consideramos amigos. Muchas veces nos ven como hijos. Y cada vez que llegamos, ah, bueno, llegó mi hijo. Así se trata, ¿no? Como una familia"'],

    ["D4: RELACIÓN\nBURÓCRATA-ADULTO MAYOR", "Nivel de escucha activa: ALTO", "SÍ (Alta)",
     '"Ver también qué problemas tienen, por ejemplo, los abuelitos, el tema psicológico... Vemos si sufre algún maltrato... Ver en qué situación viven"'],

    ["D4: RELACIÓN\nBURÓCRATA-ADULTO MAYOR", "Reconocimiento de autonomía", "SÍ (Alta)",
     '"Usted tiene la fuerza de elegir, entonces entre ellos también se designan... Para que puedan tener una decisión de voto entre su comunidad"'],

    ["D4: RELACIÓN\nBURÓCRATA-ADULTO MAYOR", "Construcción del adulto mayor: Familia + Amigo", "SÍ (Alta)",
     '"Nos consideramos amigos... Como una familia... Para el tema financiero, un poquito comentarles el tema en qué deben invertir"'],
]

# Escribir datos
for row_idx, row_data in enumerate(data, start=6):
    for col_idx, cell_value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Ajustar altura de filas
for row in range(6, 6 + len(data)):
    ws.row_dimensions[row].height = 60

# Agregar hoja de resumen
ws2 = wb.create_sheet("Resumen")

ws2['A1'] = "SÍNTESIS DEL ANÁLISIS"
ws2['A1'].font = Font(bold=True, size=14, color="1F4E78")

resumen_data = [
    ["", ""],
    ["Dimensión", "Síntesis"],
    ["D1: Discrecionalidad", "PRESENTE (Intensidad: ALTA) - Jhonatan ejerce discrecionalidad sustancial y reflexiva. Adapta normas formales según realidad local (DNI faltante, acceso territorial, vulnerabilidad)"],
    ["D2: Rutinización", "PRESENTE (Intensidad: ALTA) - Ha desarrollado sistema altamente rutinizado: delegados, protocolo de visita, WhatsApp. Paradoja: usa rutinización para MULTIPLICAR capacidad de discrecionalidad"],
    ["D3: Racionamiento", "PRESENTE (Intensidad: ALTA) - Racionamiento indirecto y sofisticado. No rechaza, pero gradúa velocidad y profundidad de atención según complejidad y vulnerabilidad"],
    ["D4: Relación", "PRESENTE (Intensidad: ALTA) - Relación horizontal, familiar, de amigos. No trato vertical burocrático. Legitima sus decisiones discrecionales a través de confianza"],
    ["", ""],
    ["CONCLUSIÓN", "Jhonatan Dávila es un burócrata de calle ALTAMENTE DISCRECIONAL que opera adaptando política nacional a realidades territoriales imposibles de estandarizar desde arriba"],
]

for row_idx, row_data in enumerate(resumen_data, start=1):
    for col_idx, cell_value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=cell_value)
        cell.border = border

        if row_idx == 2:  # Encabezado
            cell.fill = header_fill
            cell.font = header_font
        elif row_idx in [3, 4, 5, 6]:  # Dimensiones
            if col_idx == 1:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.font = Font(bold=True)

        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 120

for row in range(3, 7):
    ws2.row_dimensions[row].height = 60

# Guardar
output_path = Path("~/Desktop/Analisis_Proteccion_Social/resultados/Analisis_Cualitativo_Jhonatan.xlsx").expanduser()
wb.save(str(output_path))

print(f"✓ Archivo Excel creado: {output_path.name}")
print(f"  Tamaño: {output_path.stat().st_size / 1024:.1f} KB")
